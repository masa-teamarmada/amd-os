# -*- coding: utf-8 -*-
"""SX_コスト試算_260820 の ③詳細費用積上げ を project_cost_* の seed SQL へ変換する。"""
import openpyxl, re, json

WB = '/private/tmp/claude-501/-Users-masa-projects-AMD-SX/a6a11207-32d3-45cb-8952-6192901152a4/scratchpad/sx_cost.xlsx'
wbv = openpyxl.load_workbook(WB, data_only=True)
wbf = openpyxl.load_workbook(WB, data_only=False)
wv, wf = wbv['③詳細費用積上げ'], wbf['③詳細費用積上げ']

MODEL = 'cm_p21_260820'

def q(v):
    if v is None or v == '':
        return 'null'
    return "'" + str(v).replace("'", "''") + "'"

def n(v):
    if v is None or v == '':
        return 'null'
    try:
        return repr(round(float(v), 8))
    except (TypeError, ValueError):
        return 'null'

out = []
out.append("BEGIN;")
out.append(f"delete from project_cost_questions where cost_model_id = {q(MODEL)};")
out.append(f"delete from project_cost_items where cost_model_id = {q(MODEL)};")
out.append(f"delete from project_cost_assumptions where cost_model_id = {q(MODEL)};")
out.append(f"delete from project_cost_models where cost_model_id = {q(MODEL)};")

# ---------- model ----------
summary = (
 "ちこ作成の260820版をAMD OSへ移した初版。\\n\\n"
 "**ケースの注意**: 260820スプレッドシートには「色素」「脱色」「染料」の語が1つも無く、"
 "必要菌体量が `金属濃度 ÷ 取り込み効率α(g金属/g菌体)` で回っていた。"
 "まさの確認では今の対象は色素分解。設備構成(A循環/B投入/C中央培養)は両ケース共通なので、"
 "設備明細はそのまま使えるが、濃度・取り込み効率・酸処理まわりは色素の条件へ差し替えが要る。"
 "この版はケースを色素分解として登録し、差し替え待ちの行を要確認として残している。"
)
out.append(
 "insert into project_cost_models (cost_model_id, project_id, title, case_kind, case_label, "
 "version_label, status, source_url, source_note, summary_md, visibility, created_by, updated_by) values ("
 f"{q(MODEL)}, 'p21', 'SX コスト試算', 'dye_degradation', '色素分解', '260820版', 'active', "
 "'https://docs.google.com/spreadsheets/d/1ZGYjuQ-7grss83gVrVZ7jV7zd76AVuWDPqwi-JDjmVw/edit', "
 "'原典: Google Sheets SX_コスト試算_260820 (ちこ)。中島先生の設備・運用想定を反映した版。', "
 f"{q(summary)}, 'amd_internal', 'amie', 'amie');"
)

# ---------- assumptions ----------
# (role_key, group, label, cell, unit, confidence, source, owner, is_key, note_cell_or_text)
A = [
 ('sale_price','共通前提','想定売上単価','B5','円/m³','C','仮置き','AMD',True,
  '既存処理コスト500円/m³前後という前提から置いた値。用途別の実勢価格では未検証。'),
 ('daily_volume','共通前提','日処理量','B6','m³/日','B','先生回答','中島先生',False,'260812打合せ：m³/日基準'),
 ('operating_days','共通前提','稼働日/年','B7','日/年','B','仮置き','AMD',False,None),
 ('batch_volume','共通前提','バッチ容量','B8','m³/バッチ','B','先生回答','中島先生',False,None),
 ('utilization','共通前提','稼働率','B9','-','C','仮置き','AMD',False,'仮置き'),
 ('labor_rate','共通前提','人件費単価','B12','円/h','B','仮置き','AMD',False,None),
 ('tank_life_years','槽費用','槽の償却年数','B21','年','B','先生回答','中島先生',False,'出所：中島試算シート'),
 ('new_tank_capex','槽費用','新設槽CAPEX（コンクリート地下タンク100m³）','B19','円','B','先生回答','中島先生',False,'出所：中島試算シート'),
 ('target_concentration','対象物質','対象物質の濃度','H6','ppm','A','仮置き','杉浦先生',True,
  '★V-401。260820版では「金属濃度」。色素分解ケースでは色素濃度（mg/L等）へ差し替えが要る。'),
 ('uptake_alpha','対象物質','取り込み効率α','H7','g/g菌体','H','仮置き','杉浦先生',True,
  '★V-402。260820版の単位は g金属/g菌体。色素分解ケースでは g色素/g菌体 の吸着容量へ差し替えが要る。最重要仮説。'),
 ('recovery_eta','対象物質','菌体回収率η','H8','%','H','仮置き','中島先生',False,'V-501。00_変数辞書 Mid/プロセス分解P2。'),
 ('cell_density','対象物質','運転時菌体濃度','H9','g-DCW/L','A','先生回答','杉浦先生',False,'プロセス分解P1。5g/L。'),
 ('k_ppm','対象物質','ppm→g/m³ 換算 k_ppm','H10','g/m³/ppm','S','実測','AMD',False,'固定値'),
 ('reuse_count','対象物質','菌体使用回数（1＝使い捨て）','H11','回','H','要検証','杉浦先生',True,
  '★最大の感度。1＝再利用せず毎回新品。技術的事実ではない保守的仮定。'
  'まさの記憶では色素分解なら数十回使い回せる可能性があるが未確認（2026-08-23時点でうろ覚えと明言）。'
  '10回を超えると中央培養コストは誤差になる。MTGで判定する。'),
 ('supply_sites','中央培養','供給先数（中央培養の按分先）','G35','拠点','C','仮置き','AMD',False,
  '中央培養CAPEXの配賦先数。1＝1社で全量負担（最も保守的）。'),
 ('hrt_circulation','循環方式','HRT（反応時間・基準ケース）','B26','h','H','推定','中島先生',False,'AI比較資料をもとに暫定4h。1/4/8hで感度をみる。'),
 ('batch_hours_circulation','循環方式','1バッチあたり作業時間','B27','h','A','先生回答','中島先生',False,'出所：中島試算シート100m³ケース①循環'),
 ('module_durability_batches','循環方式','菌体保持モジュール耐用バッチ数','B28','回','H','要検証','中島先生',True,
  '20/50/100回で感度分析。実験で耐久性確認。現時点で確定値ではない。'),
 ('module_unit_price','循環方式','菌体保持モジュール一式単価','B40','円/回','H','予測','中島先生',True,
  '★当方予測：1モジュール50万円×カートリッジ3基＝150万円。中島先生の数字ではない。'
  '必要膜面積・カートリッジ数が決まってから積算し直す。A方式コストの約6割を占める。'),
 ('hrt_injection','投入方式','HRT（反応時間・基準ケース）','B32','h','H','推定','中島先生',False,'循環方式と同じくAI比較資料をもとに暫定4h'),
 ('batch_hours_injection','投入方式','1バッチあたり作業時間','B33','h','A','先生回答','中島先生',False,'出所：中島試算シート100m³ケース②投入'),
 ('membrane_life_years','投入方式','膜交換年数','B34','年','B','先生回答','中島先生',False,'260812打合せ：UF/MF2層構造、3年目安'),
 ('power_unit_price','電力','電力単価','B37','円/kWh','H','推定','AMD',False,'産業用電力の目安単価。契約単価が分かり次第更新'),
 ('power_kw_circulation','電力','循環方式：稼働動力','B38','kW','H','推定','AMD',False,'循環ポンプ2台×0.75kW程度と仮定'),
 ('power_kw_injection','電力','投入方式：稼働動力','B39','kW','H','推定','AMD',False,'撹拌機1.5kW＋UF/MF系ポンプ計1.0kW程度と仮定'),
]
for i, (rk, grp, label, cell, unit, conf, src, owner, key, note) in enumerate(A):
    val = wv[cell].value
    out.append(
      "insert into project_cost_assumptions (cost_assumption_id, cost_model_id, group_label, label, "
      "value, unit, confidence, source_kind, owner, is_key, role_key, note, sort_order) values ("
      f"{q('ca_'+rk)}, {q(MODEL)}, {q(grp)}, {q(label)}, {n(val)}, {q(unit)}, {q(conf)}, {q(src)}, "
      f"{q(owner)}, {'true' if key else 'false'}, {q(rk)}, {q(note)}, {i*10});"
    )
# 非数値の前提
out.append(
  "insert into project_cost_assumptions (cost_assumption_id, cost_model_id, group_label, label, "
  "value_text, confidence, source_kind, owner, role_key, note, sort_order) values ("
  f"{q('ca_tank_mode')}, {q(MODEL)}, '槽費用', '槽（既設 or 新設）', '既設', 'B', '仮置き', 'ダイキアクシス', "
  "'tank_mode', '既設槽を流用できるか、新設が要るかは顧客工場の現物次第。4シナリオの片側を決める。', 1000);"
)

# ---------- items ----------
BASIS = {'初期投資配賦':'初期投資配賦','毎m³比例':'毎m³比例','バッチ連動':'バッチ連動'}
def scen_of(a):
    a = (a or '').strip()
    if a == '中央培養': return '中央培養'
    if a == '循環': return '循環'
    if a == '投入': return '投入'
    if a.replace(' ', '') in ('循環,投入', '循環，投入'): return '共通'
    return None

def owner_of(src, note):
    t = note or ''
    if 'ダイキ' in t: return 'ダイキアクシス'
    if '杉浦' in t: return '杉浦先生'
    if src == '先生回答': return '中島先生'
    return 'AMD'

last_scen, seq = '共通', 0
for r in range(43, 180):
    a = wv.cell(r, 1).value
    b = wv.cell(r, 2).value
    if not a or not b:
        continue
    a, b = str(a).strip(), str(b).strip()
    is_bd = (a == '内訳') or (a == '参考')
    scen = scen_of(a)
    if scen is None:
        scen = last_scen
    else:
        last_scen = scen
    cost_type = b if b in ('CAPEX', 'OPEX', '参考') else '参考'
    raw_basis = str(wv.cell(r, 6).value or '')
    basis = '内訳' if is_bd or raw_basis.startswith('内訳') else BASIS.get(raw_basis, '毎m³比例')
    formula = wf.cell(r, 10).value
    price_rule, unit_price = None, wv.cell(r, 10).value
    if isinstance(formula, str) and formula.startswith('='):
        if '$H$16' in formula:
            price_rule = 'broth'
            m = re.match(r'^=([0-9.]+)\*', formula)
            unit_price = float(m.group(1)) if m else unit_price
        elif '$H$15' in formula:
            price_rule = 'biomass'
            m = re.match(r'^=([0-9.]+)\*', formula)
            unit_price = float(m.group(1)) if m else unit_price
        elif '$B$40' in formula:
            price_rule, unit_price = 'module_swap', 0
        elif '$B$38' in formula:
            price_rule, unit_price = 'power_circulation', 0
        elif '$B$39' in formula:
            price_rule, unit_price = 'power_injection', 0
    src = wv.cell(r, 17).value
    src = None if src in (None, '', '－') else str(src).strip()
    conf = wv.cell(r, 18).value
    conf = None if conf in (None, '', '－') else str(conf).strip()
    note = wv.cell(r, 20).value
    seq += 10
    out.append(
      "insert into project_cost_items (cost_item_id, cost_model_id, scenario, cost_type, group_label, "
      "mid_label, leaf_label, basis, quantity, quantity_unit, unit_price, unit_price_unit, price_rule, "
      "annual_factor, useful_life_years, is_breakdown, confidence, source_kind, owner, note, sort_order) values ("
      f"{q('ci_260820_%03d' % r)}, {q(MODEL)}, {q(scen)}, {q(cost_type)}, {q(wv.cell(r,3).value)}, "
      f"{q(wv.cell(r,4).value)}, {q(wv.cell(r,5).value)}, {q(basis)}, {n(wv.cell(r,8).value) if wv.cell(r,8).value is not None else '1'}, "
      f"{q(wv.cell(r,9).value)}, {n(unit_price) if unit_price is not None else '0'}, {q(wv.cell(r,11).value)}, "
      f"{q(price_rule)}, {n(wv.cell(r,12).value) if wv.cell(r,12).value is not None else '1'}, "
      f"{n(wv.cell(r,13).value)}, {'true' if is_bd else 'false'}, {q(conf)}, {q(src)}, "
      f"{q(owner_of(src, note))}, {q(note)}, {seq});"
    )

out.append("COMMIT;")
open('/private/tmp/claude-501/-Users-masa-projects-AMD-SX/a6a11207-32d3-45cb-8952-6192901152a4/scratchpad/seed_cost.sql','w',encoding='utf-8').write('\n'.join(out))
print('statements:', len(out))
